import openpyxl
import json
import sys

def analyze_excel(file_path):
    """Analyze Excel file and extract structure, formulas, and data"""
    wb = openpyxl.load_workbook(file_path, data_only=False)  # data_only=False preserves formulas
    
    result = {
        'sheet_names': wb.sheetnames,
        'sheets': {}
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'headers': [],
            'formulas': [],
            'sample_data': []
        }
        
        # Get headers (first row)
        if ws.max_row > 0:
            headers = []
            for col in range(1, min(ws.max_column + 1, 50)):  # Limit to first 50 columns
                cell = ws.cell(row=1, column=col)
                if cell.value:
                    header_info = {
                        'column': col,
                        'header': str(cell.value)
                    }
                    if cell.data_type == 'f':  # Formula
                        header_info['formula'] = cell.value  # Formula as string
                    headers.append(header_info)
            sheet_data['headers'] = headers
        
        # Find formulas
        formula_cells = []
        for row in range(1, min(ws.max_row + 1, 100)):  # Check first 100 rows
            for col in range(1, min(ws.max_column + 1, 50)):
                cell = ws.cell(row=row, column=col)
                if cell.data_type == 'f':  # Formula
                    formula_cells.append({
                        'cell': f"{cell.coordinate}",
                        'formula': str(cell.value),  # Formula as string
                        'calculated_value': str(cell.value) if cell.value else None
                    })
        sheet_data['formulas'] = formula_cells[:50]  # Limit to first 50 formulas
        
        # Get sample data (first few rows)
        sample_rows = []
        for row_idx in range(1, min(4, ws.max_row + 1)):
            row_data = {}
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    row_data[f"Col{col_idx}"] = str(cell.value)[:100]  # Limit length
            if row_data:
                sample_rows.append(row_data)
        sheet_data['sample_data'] = sample_rows
        
        result['sheets'][sheet_name] = sheet_data
    
    return result

if __name__ == "__main__":
    import os
    # Get the directory where this script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, "Stash Auto Lookup.xlsx")
    try:
        analysis = analyze_excel(file_path)
        print(json.dumps(analysis, indent=2))
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

